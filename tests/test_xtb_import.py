"""XTB xlsx import: parser (volumes, partial fills, Excel serial dates,
transfers, amount validation), duplicate detection via external_id, and the
preview/commit API flow."""
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

import db
import web.server as server
from importers.xtb import map_ticker, parse_xtb_export, rate_is_unity

REAL_EXPORT = Path("/Users/mateusz/Downloads/54998934/export_nowy.xlsx")
REAL_EXPORT_PLN = Path("/Users/mateusz/Downloads/50906523/PLN_50906523_2006-01-01_2026-07-04.xlsx")


def build_xtb_xlsx(cash_rows, closed_rows=()):
    """An in-memory workbook shaped exactly like XTB's export: metadata junk
    above the header, a Total row at the bottom, both sheets present."""
    wb = openpyxl.Workbook()
    closed = wb.active
    closed.title = "Closed Positions"
    closed.append(["Account", "12345678"])
    closed.append(["Closed Positions", ""])
    closed.append(["Instrument", "Category", "Ticker", "Type", "Volume", "Open Price"])
    for row in closed_rows:
        closed.append(row)
    closed.append(["Profit/loss"])

    cash = wb.create_sheet("Cash Operations")
    cash.append(["Account number", "12345678"])
    cash.append(["Cash Operations", ""])
    cash.append(["Date from (UTC)", datetime(2025, 9, 30, 22)])
    cash.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment", "Product"])
    for row in cash_rows:
        cash.append(row)
    cash.append(["Total", None, None, None, 1.37])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


SAMPLE_ROWS = [
    ["Stock purchase", "EGLN.UK", "Physical Gold", datetime(2026, 6, 1, 10, 36),
     "-67.12", "1288170641", "OPEN BUY 0.8942 @ 75.0600", "My Trades"],
    # partial fill: volume of THIS operation is before the slash
    ["Stock purchase", "VWCE.DE", "FTSE All-World", datetime(2026, 5, 29, 7, 5),
     -37.26, "1284805510", "OPEN BUY 0.228/1.228 @ 163.40", "My Trades"],
    # Excel serial date instead of a datetime (46174 = 2026-06-01)
    ["Stock purchase", "AAPL.US", "Apple", 46174.44,
     -371.0, "1284805999", "OPEN BUY 2 @ 185.50", "My Trades"],
    # both sale spellings appear in real exports
    ["Stock sale", "AAPL.US", "Apple", datetime(2026, 6, 2, 15, 0),
     190.0, "1290000001", "CLOSE BUY 1 @ 190.00", "My Trades"],
    ["Stock sell", "AAPL.US", "Apple", datetime(2026, 6, 3, 15, 0),
     95.5, "1290000002", "CLOSE BUY 0.5/1 @ 191.00", "My Trades"],
    ["Deposit", "", "", datetime(2026, 5, 28, 17, 16),
     492.84, "1284343298", "Pekao S.A. deposit, id=33282295", "My Trades"],
    ["Transfer", "", "", datetime(2026, 6, 1, 10, 35),
     126.99, "1288169899", "Currency conversion, PLN to EUR from TA: 1 to: 2", "My Trades"],
    # amount ≉ shares×price → implied FX rate ~1.4 (verdict belongs to preview)
    ["Stock purchase", "SXR8.DE", "Core S&P 500", datetime(2026, 5, 29, 7, 6),
     -100.0, "1284806253", "OPEN BUY 0.1021 @ 699.50", "My Trades"],
    ["Dividend", "KO.US", "Coca-Cola", datetime(2026, 5, 20, 12, 0),
     1.23, "1280000000", "KO.US USD 0.485/ SHR", "My Trades"],
    ["Withholding tax", "KO.US", "Coca-Cola", datetime(2026, 5, 20, 12, 0),
     -0.19, "1280000001", "KO.US USD WHT 15%", "My Trades"],
    ["Free funds interest", "", "", datetime(2026, 6, 3, 6, 0),
     0.31, "1290000010", "Free-funds Interest 2026-05", "My Trades"],
    ["Free funds interest tax", "", "", datetime(2026, 6, 3, 6, 0),
     -0.06, "1290000011", "Free-funds Interest Tax 2026-05", "My Trades"],
    # outgoing conversion (negative) — direction must come from the sign
    ["Transfer", "", "", datetime(2026, 6, 4, 10, 0),
     -540.42, "1290000012", "Currency conversion, PLN to EUR from TA: 1 to: 2", "My Trades"],
    ["Withdrawal", "", "", datetime(2026, 6, 5, 10, 0),
     -64.85, "1290000013", "Withdrawal from 50906523", "My Trades"],
    ["Subaccount transfer", "", "", datetime(2026, 3, 5, 10, 0),
     508.21, "1290000014", "Transfer from 54388868 to 50906523", "My Trades"],
    ["Rights issue", "ORSTED.DK", "Orsted", datetime(2025, 9, 18, 10, 0),
     18.69, "1290000015", "ORSTED.DK DKK 93.1800/ SHR", "My Trades"],
    ["Dividend zero", "", "", datetime(2026, 5, 20, 12, 0),
     1.0, "1290000016", "unsupported type row", "My Trades"],
]

# ops the parser should emit from SAMPLE_ROWS (the "Dividend zero" row is an
# unsupported type and only produces a file-level warning)
SAMPLE_OPS = 16


# ---- Parser -----------------------------------------------------------------

def test_parser_on_synthetic_export():
    result = parse_xtb_export(build_xtb_xlsx(SAMPLE_ROWS))
    ops = result["operations"]
    assert len(ops) == SAMPLE_OPS
    by_kind = {}
    for op in ops:
        by_kind.setdefault(op["kind"], []).append(op)

    assert len(by_kind["BUY"]) == 4
    assert len(by_kind["SELL"]) == 2  # "Stock sale" + "Stock sell"
    assert len(by_kind["DEPOSIT"]) == 1
    assert len(by_kind["TRANSFER"]) == 2
    assert len(by_kind["DIVIDEND"]) == 1
    assert len(by_kind["TAX"]) == 2  # WHT + interest tax
    assert len(by_kind["INTEREST"]) == 1
    assert len(by_kind["WITHDRAWAL"]) == 1
    assert len(by_kind["SUBTRANSFER"]) == 1
    assert len(by_kind["RIGHTS"]) == 1

    egln = next(o for o in ops if o["external_id"] == "1288170641")
    assert egln["ticker"] == "EGLN.L"  # .UK → .L
    assert egln["shares"] == pytest.approx(0.8942)
    assert egln["price"] == pytest.approx(75.06)
    assert egln["amount"] == pytest.approx(67.12)  # string amount coerced
    assert rate_is_unity(egln["implied_rate"], egln["amount"])

    partial = next(o for o in ops if o["external_id"] == "1284805510")
    assert partial["shares"] == pytest.approx(0.228)  # NOT 1.228

    serial = next(o for o in ops if o["external_id"] == "1284805999")
    assert serial["date"] == "2026-06-01"  # 46174.44 days since 1899-12-30
    assert serial["ticker"] == "AAPL"  # .US → no suffix

    sales = by_kind["SELL"]
    assert (sales[0]["shares"], sales[0]["price"]) == (1, 190.0)
    assert (sales[1]["shares"], sales[1]["price"]) == (0.5, 191.0)  # partial close

    # implied FX rate reported, judged by the preview (needs instrument currency)
    fx_buy = next(o for o in ops if o["external_id"] == "1284806253")
    assert fx_buy["implied_rate"] == pytest.approx(1.4, abs=0.01)
    assert not rate_is_unity(fx_buy["implied_rate"], fx_buy["amount"])
    assert fx_buy["warning"] is None

    # unsupported type row → file-level warning, not an operation
    assert all(o["external_id"] != "1290000016" for o in ops)
    assert any("Dividend zero" in w for w in result["warnings"])


def test_cash_flows_direction_and_notes():
    ops = parse_xtb_export(build_xtb_xlsx(SAMPLE_ROWS))["operations"]
    by_id = {o["external_id"]: o for o in ops}

    transfer_in = by_id["1288169899"]
    assert (transfer_in["kind"], transfer_in["cash_type"]) == ("TRANSFER", "DEPOSIT")
    assert transfer_in["note"] == "przewalutowanie PLN→EUR"

    transfer_out = by_id["1290000012"]  # negative amount → outflow
    assert (transfer_out["kind"], transfer_out["cash_type"]) == ("TRANSFER", "WITHDRAWAL")
    assert transfer_out["amount"] == pytest.approx(540.42)

    dividend = by_id["1280000000"]
    assert (dividend["kind"], dividend["cash_type"]) == ("DIVIDEND", "DEPOSIT")
    assert dividend["note"] == "dywidenda KO"
    assert dividend["ticker"] == "KO"  # display only

    wht = by_id["1280000001"]
    assert (wht["kind"], wht["cash_type"]) == ("TAX", "WITHDRAWAL")
    assert wht["note"] == "podatek u źródła KO"

    interest = by_id["1290000010"]
    assert (interest["kind"], interest["cash_type"]) == ("INTEREST", "DEPOSIT")
    assert interest["note"] == "odsetki od wolnych środków 2026-05"

    interest_tax = by_id["1290000011"]
    assert (interest_tax["kind"], interest_tax["cash_type"]) == ("TAX", "WITHDRAWAL")
    assert interest_tax["note"] == "podatek od odsetek 2026-05"

    withdrawal = by_id["1290000013"]
    assert (withdrawal["kind"], withdrawal["cash_type"]) == ("WITHDRAWAL", "WITHDRAWAL")

    sub = by_id["1290000014"]
    assert (sub["kind"], sub["cash_type"]) == ("SUBTRANSFER", "DEPOSIT")
    assert sub["note"] == "transfer między subkontami"

    rights = by_id["1290000015"]
    assert (rights["kind"], rights["cash_type"]) == ("RIGHTS", "DEPOSIT")
    assert rights["note"] == "prawa poboru ORSTED.CO"  # .DK → .CO


def test_amount_within_tolerance_counts_as_unity():
    rows = [["Stock purchase", "VWCE.DE", "x", datetime(2026, 1, 2),
             -163.9, "1", "OPEN BUY 1 @ 163.48", "My Trades"]]  # 0.26% off
    op = parse_xtb_export(build_xtb_xlsx(rows))["operations"][0]
    assert rate_is_unity(op["implied_rate"], op["amount"])


def test_closed_positions_with_rows_warns():
    closed = [["FTSE All-World", "ETF", "VWCE.DE", "BUY", 1.0, 100.0]]
    result = parse_xtb_export(build_xtb_xlsx(SAMPLE_ROWS[:1], closed_rows=closed))
    assert any("Closed Positions" in w for w in result["warnings"])
    assert len(result["operations"]) == 1  # nothing duplicated from that sheet


def test_not_an_xtb_file_raises():
    wb = openpyxl.Workbook()
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Cash Operations"):
        parse_xtb_export(buf.getvalue())
    with pytest.raises(ValueError):
        parse_xtb_export(b"definitely not xlsx")


def test_suffix_mapping():
    assert map_ticker("EGLN.UK") == "EGLN.L"
    assert map_ticker("CDR.PL") == "CDR.WA"
    assert map_ticker("AAPL.US") == "AAPL"
    assert map_ticker("SXR8.DE") == "SXR8.DE"
    assert map_ticker("TTE.FR") == "TTE.PA"
    assert map_ticker("ASML.NL") == "ASML.AS"
    assert map_ticker("ENI.IT") == "ENI.MI"
    assert map_ticker("SAN.ES") == "SAN.MC"
    assert map_ticker("WEIRD.XX") == "WEIRD.XX"  # unknown suffix passes through


@pytest.mark.skipif(not REAL_EXPORT.exists(), reason="real XTB export not present")
def test_parser_on_real_export():
    result = parse_xtb_export(REAL_EXPORT.read_bytes())
    ops = result["operations"]
    kinds = [o["kind"] for o in ops]
    assert kinds.count("BUY") == 18
    assert kinds.count("DEPOSIT") == 4
    assert kinds.count("TRANSFER") == 3
    assert result["warnings"] == []
    # EUR account, EUR/EUR-quoted instruments: every amount reconciles at rate 1
    assert all(rate_is_unity(o["implied_rate"], o["amount"]) for o in ops if o["kind"] == "BUY")
    partial = next(o for o in ops if o["external_id"] == "1284805510")
    assert partial["shares"] == pytest.approx(0.228)


@pytest.mark.skipif(not REAL_EXPORT_PLN.exists(), reason="real PLN XTB export not present")
def test_parser_on_real_pln_export():
    result = parse_xtb_export(REAL_EXPORT_PLN.read_bytes())
    ops = result["operations"]
    kinds = [o["kind"] for o in ops]
    assert kinds.count("BUY") == 41
    assert kinds.count("SELL") == 30  # "Stock sell" spelling
    assert kinds.count("DEPOSIT") == 13
    assert kinds.count("WITHDRAWAL") == 3
    assert kinds.count("TRANSFER") == 3
    assert kinds.count("DIVIDEND") == 6
    assert kinds.count("TAX") == 18  # 6 WHT + 12 interest tax
    assert kinds.count("INTEREST") == 12
    assert kinds.count("RIGHTS") == 6
    assert kinds.count("SUBTRANSFER") == 4

    # outgoing PLN→EUR conversions: negative amounts → withdrawals
    assert all(o["cash_type"] == "WITHDRAWAL" for o in ops if o["kind"] == "TRANSFER")
    # foreign instruments (EUR/DKK) on the PLN account: implied rate ≠ 1
    orsted = [o for o in ops if o["kind"] == "SELL" and o["xtb_ticker"] == "ORSTED.DK"]
    assert orsted and all(not rate_is_unity(o["implied_rate"], o["amount"]) for o in orsted)
    # non-empty Closed Positions sheet is reported, not imported
    assert any("Closed Positions" in w for w in result["warnings"])


# ---- API: preview + commit ---------------------------------------------------

@pytest.fixture()
def client(monkeypatch):
    monkeypatch.setattr(server, "_cached_price", lambda ticker: 100.0)
    monkeypatch.setattr(server, "get_current_price", lambda ticker: 100.0)
    monkeypatch.setattr(server, "get_instrument_info", lambda t: {
        "ticker": t.upper(), "name": f"{t.upper()} Inc.", "type": "ETF",
        "exchange": "TEST", "currency": "EUR",
    } if t.upper() != "BOGUS.XX" else None)
    server._price_cache.clear()
    server._fx_cache.clear()
    server._history_cache.clear()
    return TestClient(server.app)


@pytest.fixture()
def portfolio_id(client):
    r = client.post("/api/portfolios", json={"name": "Import testowy", "currency": "EUR"})
    assert r.status_code == 201, r.text
    pid = r.json()["id"]
    yield pid
    client.delete(f"/api/portfolios/{pid}?force=true")


def upload(client, pid, content):
    return client.post(
        f"/api/portfolios/{pid}/import/xtb",
        files={"file": ("export.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_preview_commit_and_reimport_flow(client, portfolio_id):
    content = build_xtb_xlsx(SAMPLE_ROWS)

    preview = upload(client, portfolio_id, content)
    assert preview.status_code == 200, preview.text
    ops = preview.json()["operations"]
    assert len(ops) == SAMPLE_OPS
    assert all(op["already_exists"] is False for op in ops)
    trades = [op for op in ops if op["kind"] in ("BUY", "SELL")]
    assert all(op["ticker_verified"] is True for op in trades)
    # cash flows carry a ticker only for display — never verified
    assert all(op["ticker_verified"] is None for op in ops if op not in trades)

    # instrument currency == portfolio currency (EUR fixture) + rate 1.4 →
    # a real mismatch warning; the same-currency rows stay clean
    fx_buy = next(op for op in ops if op["external_id"] == "1284806253")
    assert fx_buy["warning"] and "nie zgadza się" in fx_buy["warning"]
    assert all(op["warning"] is None for op in trades if op is not fx_buy)

    commit = client.post(
        f"/api/portfolios/{portfolio_id}/import/xtb/commit",
        json={"operations": ops},
    )
    assert commit.status_code == 201, commit.text
    assert commit.json() == {"imported": SAMPLE_OPS, "skipped_duplicates": 0}

    # engine state after import: 2 AAPL bought, 1 + 0.5 sold
    summary = client.get(f"/api/portfolios/{portfolio_id}/summary").json()
    tickers = {p["ticker"]: p for p in summary["positions"]}
    assert tickers["AAPL"]["shares"] == pytest.approx(0.5)
    assert tickers["EGLN.L"]["shares"] == pytest.approx(0.8942)

    deposits = client.get(f"/api/portfolios/{portfolio_id}/deposits").json()
    assert len(deposits) == 10  # every non-trade kind lands in deposits
    flows = {(d["note"], d["type"]) for d in deposits}
    # both conversion directions share the note — the type tells them apart
    assert ("przewalutowanie PLN→EUR", "DEPOSIT") in flows
    assert ("przewalutowanie PLN→EUR", "WITHDRAWAL") in flows
    assert ("dywidenda KO", "DEPOSIT") in flows
    assert ("podatek u źródła KO", "WITHDRAWAL") in flows
    assert ("odsetki od wolnych środków 2026-05", "DEPOSIT") in flows
    assert ("podatek od odsetek 2026-05", "WITHDRAWAL") in flows
    summary2 = client.get(f"/api/portfolios/{portfolio_id}/summary").json()
    # rights issues (RETURN) are in the P&L base (not profit) but NOT in the
    # user-facing "suma wpłat" — they're not user-deposited capital
    non_profit_base = 492.84 + 126.99 - 540.42 - 64.85 + 508.21 + 18.69
    assert summary2["deposited"] == pytest.approx(non_profit_base)
    assert summary2["contributed_in"] == pytest.approx(492.84 + 126.99 + 508.21)
    assert summary2["contributed_out"] == pytest.approx(540.42 + 64.85)
    contributed = non_profit_base
    # …while cash additionally includes investment income and uses the EXACT
    # imported trade amounts (not shares×price×current FX)
    income = 1.23 - 0.19 + 0.31 - 0.06
    trades_cash = -67.12 - 37.26 - 371.0 + 190.0 + 95.5 - 100.0
    assert summary2["cash"] == pytest.approx(contributed + income + trades_cash)
    # lifetime P&L is money-weighted: value + cash − contributed capital
    assert summary2["total_pnl"] == pytest.approx(
        summary2["cash"] + summary2["positions_value"] - contributed
    )

    # re-import of the same file: everything flagged, commit imports nothing
    preview2 = upload(client, portfolio_id, content)
    ops2 = preview2.json()["operations"]
    assert all(op["already_exists"] is True for op in ops2)
    commit2 = client.post(
        f"/api/portfolios/{portfolio_id}/import/xtb/commit",
        json={"operations": ops2},
    )
    assert commit2.json() == {"imported": 0, "skipped_duplicates": SAMPLE_OPS}
    assert len(client.get(f"/api/portfolios/{portfolio_id}/deposits").json()) == 10


def test_unverified_ticker_flagged_in_preview(client, portfolio_id):
    rows = [["Stock purchase", "BOGUS.XX", "Mystery", datetime(2026, 1, 2),
             -100.0, "42", "OPEN BUY 1 @ 100.00", "My Trades"]]
    ops = upload(client, portfolio_id, build_xtb_xlsx(rows)).json()["operations"]
    assert ops[0]["ticker_verified"] is False


def test_commit_rejects_uncovered_sell(client, portfolio_id):
    ops = [{"kind": "SELL", "ticker": "AAPL", "date": "2026-01-02",
            "shares": 1, "price": 100.0, "amount": 100.0,
            "external_id": "777", "note": None}]
    r = client.post(f"/api/portfolios/{portfolio_id}/import/xtb/commit",
                    json={"operations": ops})
    assert r.status_code == 400
    assert "AAPL" in r.json()["detail"]
    # nothing was written
    assert client.get(f"/api/portfolios/{portfolio_id}/summary").json()["positions"] == []


def test_commit_imported_transactions_use_portfolio_currency(client, portfolio_id):
    rows = [["Stock purchase", "EGLN.UK", "Physical Gold", datetime(2026, 6, 1),
             "-67.12", "1288170641", "OPEN BUY 0.8942 @ 75.0600", "My Trades"]]
    ops = upload(client, portfolio_id, build_xtb_xlsx(rows)).json()["operations"]
    client.post(f"/api/portfolios/{portfolio_id}/import/xtb/commit",
                json={"operations": ops})
    txns = db.get_transactions(portfolio_id, "EGLN.L")
    # amount ≈ shares×price → the price was paid in the account currency;
    # the file's price stays as-is (never rescaled, also for GBp listings)
    assert txns[0]["price"] == pytest.approx(75.06)
    assert txns[0]["currency"] == "EUR"
    assert txns[0]["external_id"] == "1288170641"


def test_preview_flags_probable_manual_duplicates(client, portfolio_id):
    """Transactions typed in by hand have no external_id — the preview must
    still spot them by ticker+date+shares+price, or importing the broker's
    full history doubles every manually-entered position."""
    client.post(f"/api/portfolios/{portfolio_id}/deposits",
                json={"amount": 1000, "date": "2026-05-01"})
    client.post(f"/api/portfolios/{portfolio_id}/transactions", json={
        "ticker": "VWCE.DE", "type": "BUY", "shares": 0.228, "price": 163.40,
        "date": "2026-05-29",
    })
    ops = upload(client, portfolio_id, build_xtb_xlsx(SAMPLE_ROWS)).json()["operations"]
    partial = next(o for o in ops if o["external_id"] == "1284805510")
    assert partial["similar_exists"] is True  # same ticker/date/shares/price
    other = next(o for o in ops if o["external_id"] == "1288170641")
    assert other["similar_exists"] is False


def test_closed_imported_foreign_position_needs_no_fx_note(client):
    """A fully-closed foreign position built entirely from imported trades is
    FX-exact (broker-settled amounts) — the summary must NOT register a
    current FX rate for it, or the UI shows a bogus '≈ przybliżenie' note."""
    r = client.post("/api/portfolios", json={"name": "FX note test", "currency": "PLN"})
    pid = r.json()["id"]
    try:
        rows = [
            ["Stock purchase", "SXRV.DE", "NASDAQ 100", datetime(2026, 3, 2),
             -247.29, "801", "OPEN BUY 0.0483 @ 1204.4", "My Trades"],
            ["Stock sell", "SXRV.DE", "NASDAQ 100", datetime(2026, 3, 5),
             254.47, "802", "CLOSE BUY 0.0483 @ 1236.2", "My Trades"],
        ]
        ops = upload(client, pid, build_xtb_xlsx(rows)).json()["operations"]
        client.post(f"/api/portfolios/{pid}/import/xtb/commit", json={"operations": ops})
        s = client.get(f"/api/portfolios/{pid}/summary").json()
        [closed] = s["closed_positions"]
        assert closed["fx_exact"] is True
        assert closed["realized_pnl_pc"] == pytest.approx(254.47 - 247.29)
        assert s["fx_rates"] == {}  # nothing was converted at the current rate
        assert s["fx_unavailable"] == []
    finally:
        client.delete(f"/api/portfolios/{pid}?force=true")


def test_commit_foreign_instrument_keeps_instrument_currency(client):
    """EUR-quoted instrument bought on a PLN account: amount = shares×price×FX,
    so the transaction must be recorded in the instrument's currency (EUR) —
    recording 1204.40 as PLN would wreck the cost basis."""
    r = client.post("/api/portfolios", json={"name": "Import PLN", "currency": "PLN"})
    pid = r.json()["id"]
    try:
        rows = [["Stock purchase", "SXRV.DE", "NASDAQ 100", datetime(2026, 3, 2),
                 -247.29, "77", "OPEN BUY 0.0483 @ 1204.4", "My Trades"]]  # rate ~4.25
        ops = upload(client, pid, build_xtb_xlsx(rows)).json()["operations"]
        # known instrument currency (EUR fixture) ≠ PLN explains the rate: no warning
        assert ops[0]["warning"] is None
        client.post(f"/api/portfolios/{pid}/import/xtb/commit", json={"operations": ops})
        txn = db.get_transactions(pid, "SXRV.DE")[0]
        assert txn["price"] == pytest.approx(1204.4)
        assert txn["currency"] == "EUR"
    finally:
        client.delete(f"/api/portfolios/{pid}?force=true")
