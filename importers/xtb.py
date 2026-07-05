"""Parser for XTB account exports (xlsx) — pure like engine.py: bytes in,
operation dicts out, no DB and no network. The web layer decides what to do
with the result (duplicate detection, ticker verification, persisting).

XTB's "Cash Operations" sheet is the source of truth: stock purchases/sales
appear there alongside every cash flow (deposits, withdrawals, currency
transfers, dividends + withholding tax, free-funds interest + its tax, rights
issues, subaccount transfers), with trade volume and price embedded in the
Comment column ("OPEN BUY 0.8942 @ 75.0600"). Comment prices are in the
INSTRUMENT's currency while Amount is in the ACCOUNT's currency — the
`implied_rate` (|Amount| / shares×price) is 1.0 for same-currency instruments
and the FX rate for foreign ones. XTB prices are what was actually paid and
are never rescaled here (notably: no GBp/GBP pence games for LSE listings).
"""
from __future__ import annotations

import re
from datetime import datetime, timedelta
from io import BytesIO

# XTB ticker suffix → Yahoo suffix. ".US" maps to no suffix at all.
SUFFIX_MAP = {
    ".UK": ".L",
    ".PL": ".WA",
    ".US": "",
    ".DE": ".DE",
    ".FR": ".PA",
    ".NL": ".AS",
    ".IT": ".MI",
    ".ES": ".MC",
    ".DK": ".CO",
}

# Cash-operation Type → semantic kind shown in the preview. The DIRECTION is
# never taken from the label — always from the Amount's sign (e.g. "Transfer"
# rows are outgoing PLN→EUR conversions on the PLN account and incoming ones
# on the EUR account).
CASH_KINDS = {
    "Deposit": "DEPOSIT",
    "Withdrawal": "WITHDRAWAL",
    "Transfer": "TRANSFER",
    "Subaccount transfer": "SUBTRANSFER",
    "Dividend": "DIVIDEND",
    "Withholding tax": "TAX",
    "Free funds interest": "INTEREST",
    "Free funds interest tax": "TAX",
    "Rights issue": "RIGHTS",
}

CASH_SHEET = "Cash Operations"
CLOSED_SHEET = "Closed Positions"
CASH_HEADER = ("Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment", "Product")

# "OPEN BUY 0.8942 @ 75.0600" / "CLOSE BUY 2 @ 163.40" — and partial fills
# "OPEN BUY 0.228/1.228 @ 163.40" where the volume of THIS operation is the
# number before the slash (the total position size follows it).
_COMMENT_RE = re.compile(
    r"(?:OPEN|CLOSE)\s+(?:BUY|SELL)\s+([\d.,]+)(?:/[\d.,]+)?\s*@\s*([\d.,]+)"
)

# Comment prices are in the INSTRUMENT's currency while Amount is in the
# ACCOUNT's currency, so |Amount| / (shares×price) is 1.0 for same-currency
# instruments and the FX rate for foreign ones (e.g. ~4.26 for EUR on a PLN
# account). The parser only reports that implied rate; judging whether it is
# legitimate needs the instrument currency, which the web layer knows.
RATE_UNITY_TOL = 0.015

_EXCEL_EPOCH = datetime(1899, 12, 30)

# "Free-funds Interest 2026-05" → the month; "Currency conversion, PLN to EUR…"
_MONTH_RE = re.compile(r"(\d{4}-\d{2})\s*$")
_CONVERSION_RE = re.compile(r"\b([A-Z]{3}) to ([A-Z]{3})\b")


def _cash_note(op_type: str, comment: str, ticker: str | None) -> str | None:
    """Polish note stored with the cash flow, so the deposits list explains
    where the money came from (dividend of what, which month's interest…)."""
    if op_type == "Transfer":
        m = _CONVERSION_RE.search(comment)
        return f"przewalutowanie {m.group(1)}→{m.group(2)}" if m else "przewalutowanie"
    if op_type == "Subaccount transfer":
        return "transfer między subkontami"
    if op_type == "Dividend":
        return f"dywidenda {ticker}" if ticker else "dywidenda"
    if op_type == "Withholding tax":
        return f"podatek u źródła {ticker}" if ticker else "podatek u źródła"
    if op_type in ("Free funds interest", "Free funds interest tax"):
        base = "odsetki od wolnych środków" if op_type == "Free funds interest" else "podatek od odsetek"
        m = _MONTH_RE.search(comment)
        return f"{base} {m.group(1)}" if m else base
    if op_type == "Rights issue":
        return f"prawa poboru {ticker}" if ticker else "prawa poboru"
    return None


def map_ticker(xtb_ticker: str) -> str:
    """XTB ticker → Yahoo ticker via the suffix table; unknown suffixes pass
    through unchanged (the preview lets the user correct them by hand)."""
    ticker = xtb_ticker.strip().upper()
    for suffix, yahoo in SUFFIX_MAP.items():
        if ticker.endswith(suffix):
            return ticker[: -len(suffix)] + yahoo
    return ticker


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def _to_iso_date(value) -> str | None:
    """XTB's Time column arrives as a datetime, an Excel serial number
    (days since 1899-12-30, e.g. 46174.44) or occasionally text."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    serial = _to_float(value)
    if serial is not None and 1 < serial < 200000:
        return (_EXCEL_EPOCH + timedelta(days=serial)).date().isoformat()
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.strip()).date().isoformat()
        except ValueError:
            return None
    return None


def _parse_comment(comment: str) -> tuple[float, float] | None:
    m = _COMMENT_RE.search(comment or "")
    if not m:
        return None
    shares = _to_float(m.group(1))
    price = _to_float(m.group(2))
    if shares is None or price is None or shares <= 0 or price < 0:
        return None
    return shares, price


def rate_is_unity(rate: float | None, amount: float) -> bool:
    """Does the implied rate say the price is in the account currency? A hair
    of absolute slack keeps sub-1-unit operations from tripping on rounding."""
    if rate is None:
        return True
    return abs(rate - 1) <= RATE_UNITY_TOL or abs(rate - 1) * abs(amount) <= 0.02


def parse_xtb_export(content: bytes) -> dict:
    """Parse an XTB xlsx export into operation dicts + file-level warnings.

    Returns {"operations": [...], "warnings": [...]}; raises ValueError when
    the file is not an XTB export (missing sheet/header). Each operation:
    kind BUY/SELL/DEPOSIT/TRANSFER, ticker (Yahoo-mapped, None for cash ops),
    xtb_ticker, date (ISO), shares/price (None for cash ops), amount (always
    positive), external_id (XTB's ID column), note, warning (row-level).
    """
    import openpyxl  # local import: keeps `import db`-style startup light

    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Nie można odczytać pliku xlsx: {exc}") from exc

    if CASH_SHEET not in wb.sheetnames:
        raise ValueError(f"Brak arkusza „{CASH_SHEET}” — to nie wygląda na eksport XTB")

    warnings: list[str] = []
    operations: list[dict] = []

    ws = wb[CASH_SHEET]
    ws.reset_dimensions()  # XTB writes broken dimension metadata (A1:A1)
    rows = ws.iter_rows(values_only=True)

    header_seen = False
    for row in rows:
        if not header_seen:
            if row and tuple(str(c or "").strip() for c in row[: len(CASH_HEADER)]) == CASH_HEADER:
                header_seen = True
            continue

        cells = list(row) + [None] * (len(CASH_HEADER) - len(row))
        op_type, ticker, _instrument, time_val, amount_val, ext_id, comment, _product = (
            cells[: len(CASH_HEADER)]
        )
        op_type = str(op_type or "").strip()
        if not op_type or op_type == "Total":
            continue

        amount = _to_float(amount_val)
        date = _to_iso_date(time_val)
        external_id = str(ext_id).strip() if ext_id not in (None, "") else None
        comment = str(comment or "").strip()

        if date is None or amount is None:
            warnings.append(f"Pominięto wiersz „{op_type}” (ID {external_id}): brak daty lub kwoty")
            continue

        xtb_ticker = str(ticker or "").strip()
        yahoo_ticker = map_ticker(xtb_ticker) if xtb_ticker else None

        if op_type in ("Stock purchase", "Stock sale", "Stock sell"):
            kind = "BUY" if op_type == "Stock purchase" else "SELL"
            parsed = _parse_comment(comment)
            if parsed is None:
                warnings.append(
                    f"Pominięto {kind} {ticker} (ID {external_id}): "
                    f"nie rozpoznano ilości/ceny w komentarzu „{comment}”"
                )
                continue
            shares, price = parsed
            operations.append({
                "kind": kind,
                "cash_type": None,
                "xtb_ticker": xtb_ticker,
                "ticker": yahoo_ticker,
                "date": date,
                "shares": shares,
                "price": price,
                "amount": abs(amount),
                # |Amount| / (shares×price): 1.0 = price in account currency,
                # anything else = the FX rate of the instrument's currency
                "implied_rate": abs(amount) / (shares * price) if shares * price > 0 else None,
                "external_id": external_id,
                "note": None,
                "warning": None,
            })
        elif op_type in CASH_KINDS:
            if amount == 0:
                warnings.append(f"Pominięto „{op_type}” (ID {external_id}): kwota 0")
                continue
            operations.append({
                "kind": CASH_KINDS[op_type],
                # direction from the sign, never from the label
                "cash_type": "DEPOSIT" if amount > 0 else "WITHDRAWAL",
                "xtb_ticker": xtb_ticker or None,
                "ticker": yahoo_ticker,  # display only (e.g. dividend's instrument)
                "date": date,
                "shares": None,
                "price": None,
                "amount": abs(amount),
                "implied_rate": None,
                "external_id": external_id,
                "note": _cash_note(op_type, comment, yahoo_ticker),
                "warning": None,
            })
        else:
            warnings.append(f"Pominięto nieobsługiwany typ operacji „{op_type}” (ID {external_id})")

    if not header_seen:
        raise ValueError(f"Nie znaleziono nagłówka w arkuszu „{CASH_SHEET}”")

    # Closed positions import via their Cash Operations rows — never twice.
    if CLOSED_SHEET in wb.sheetnames:
        closed = _count_closed_rows(wb[CLOSED_SHEET])
        if closed:
            warnings.append(
                f"Arkusz „{CLOSED_SHEET}” zawiera {closed} pozycji — zamknięte pozycje "
                f"importują się z Cash Operations, nie są duplikowane"
            )

    return {"operations": operations, "warnings": warnings}


def _count_closed_rows(ws) -> int:
    """Data rows in Closed Positions, i.e. rows after its header that carry an
    instrument (summary rows like 'Profit/loss' have nothing in column 3)."""
    ws.reset_dimensions()
    count = 0
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        first = str(row[0] or "").strip() if row else ""
        if not header_seen:
            if first == "Instrument":
                header_seen = True
            continue
        if len(row) > 2 and row[2] not in (None, ""):
            count += 1
    return count
